"""PDF and Excel export helpers for FlockIQ batch reports."""
from datetime import date
from io import BytesIO

from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

NAVY = colors.HexColor('#3d5a99')
PURPLE = colors.HexColor('#7c3aed')
STRIPE = colors.HexColor('#f4f7fe')
GRID = colors.HexColor('#e2e8f0')


def generate_credit_score_pdf(org, credit_score) -> bytes:
    """Return a PDF Farm Credit Report for an org."""
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=2 * cm,
        rightMargin=2 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
    )
    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph('FlockIQ — Farm Credit Report', styles['Title']))
    story.append(Paragraph(f'Farm: {org.name}', styles['Normal']))
    story.append(Paragraph(f'Report Date: {date.today()}', styles['Normal']))
    story.append(Spacer(1, 0.5 * cm))

    story.append(Paragraph(f'Credit Score: {credit_score.score}/100 — Grade {credit_score.grade}', styles['h2']))
    story.append(Paragraph(f'Confidence Level: {credit_score.get_confidence_display()}', styles['Normal']))
    story.append(Spacer(1, 0.4 * cm))

    sub_data = [
        ['Category', 'Score', 'Weight'],
        ['Financial Health', str(credit_score.financial_health_score), '30%'],
        ['Operational Consistency', str(credit_score.operational_consistency_score), '20%'],
        ['Mortality Management', str(credit_score.mortality_management_score), '20%'],
        ['Feed Efficiency', str(credit_score.feed_efficiency_score), '15%'],
        ['Platform Engagement', str(credit_score.platform_engagement_score), '10%'],
        ['Payment History', str(credit_score.payment_history_score), '5%'],
        ['Overall Score', str(credit_score.score), '100%'],
    ]
    t = Table(sub_data, colWidths=[8 * cm, 4 * cm, 4 * cm])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), PURPLE),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, STRIPE]),
        ('BACKGROUND', (0, -1), (-1, -1), STRIPE),
        ('GRID', (0, 0), (-1, -1), 0.5, GRID),
        ('PADDING', (0, 0), (-1, -1), 6),
    ]))
    story.append(t)
    story.append(Spacer(1, 0.5 * cm))

    stats_data = [
        ['Key Statistic', 'Value'],
        ['Batches Analysed', str(credit_score.batches_analysed)],
        ['Total Birds Managed', str(credit_score.total_birds_managed)],
        ['Avg Mortality Rate',
         f'{credit_score.avg_mortality_rate_pct:.1f}%' if credit_score.avg_mortality_rate_pct is not None else 'N/A'],
        ['Avg Profit Margin',
         f'{credit_score.avg_profit_margin_pct:.1f}%' if credit_score.avg_profit_margin_pct is not None else 'N/A'],
        ['Avg FCR',
         f'{credit_score.avg_fcr:.2f}' if credit_score.avg_fcr is not None else 'N/A'],
        ['Months on Platform', str(credit_score.months_on_platform)],
    ]
    t2 = Table(stats_data, colWidths=[8 * cm, 8 * cm])
    t2.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), NAVY),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, STRIPE]),
        ('GRID', (0, 0), (-1, -1), 0.5, GRID),
        ('PADDING', (0, 0), (-1, -1), 6),
    ]))
    story.append(t2)
    story.append(Spacer(1, 0.4 * cm))

    story.append(
        Paragraph(
            'Note: FCR calculated using surviving bird count at harvest.',
            styles['Normal'],
        )
    )
    story.append(Spacer(1, 0.4 * cm))

    story.append(
        Paragraph(
            'Verified by FlockIQ — This report was generated from verified farm management data '
            'logged on the FlockIQ platform. FlockIQ does not guarantee credit approval.',
            styles['Normal'],
        )
    )

    doc.build(story)
    return buffer.getvalue()


_LOSS_CAUSE_LABELS = {
    'disease': 'Disease',
    'heat_stress': 'Heat Stress',
    'accident': 'Accident / Equipment Failure',
    'predator': 'Predator Attack',
    'other': 'Other',
}


def _loss_report_footer(canvas, doc):
    """Draw the disclaimer footer on every page of the loss documentation report."""
    canvas.saveState()
    width = doc.pagesize[0]
    canvas.setStrokeColor(GRID)
    canvas.setLineWidth(0.5)
    canvas.line(2 * cm, 1.7 * cm, width - 2 * cm, 1.7 * cm)
    canvas.setFont('Helvetica-Bold', 7.5)
    canvas.setFillColor(NAVY)
    canvas.drawCentredString(
        width / 2, 1.4 * cm,
        'Generated by FlockIQ — verified farm management data',
    )
    canvas.setFont('Helvetica', 6.5)
    canvas.setFillColor(colors.HexColor('#8a96a8'))
    canvas.drawCentredString(
        width / 2, 1.05 * cm,
        'This report does not constitute an insurance claim or guarantee of coverage. '
        'Submit to your insurance provider as supporting documentation.',
    )
    canvas.drawCentredString(width / 2, 0.75 * cm, f'Page {doc.page}')
    canvas.restoreState()


def generate_loss_documentation_pdf(
    batch,
    loss_details,
    mortality_logs=None,
    feed_logs=None,
    vaccinations=None,
    valuation=None,
) -> bytes:
    """
    Return a PDF Loss Documentation Report for an insurance claim.

    ``loss_details`` is the cleaned_data dict from LossDocumentationForm
    (cause_of_death, incident_date, birds_affected, additional_notes).
    The remaining args are the supporting querysets/dict gathered by the view.
    """
    mortality_logs = list(mortality_logs or [])
    feed_logs = list(feed_logs or [])
    vaccinations = list(vaccinations or [])

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=2 * cm,
        rightMargin=2 * cm,
        topMargin=2 * cm,
        bottomMargin=2.3 * cm,
        title=f'Loss Documentation Report — {batch.batch_name}',
    )
    styles = getSampleStyleSheet()
    section_style = ParagraphStyle(
        'LossSection', parent=styles['h2'], textColor=NAVY, spaceBefore=6, spaceAfter=4,
    )
    subtitle_style = ParagraphStyle(
        'LossSubtitle', parent=styles['Normal'], textColor=colors.HexColor('#8a96a8'),
        alignment=TA_CENTER, fontSize=10,
    )
    title_style = ParagraphStyle(
        'LossTitle', parent=styles['Title'], textColor=NAVY, alignment=TA_CENTER,
    )
    story = []

    # ── Cover ─────────────────────────────────────────────────────────────────
    story.append(Spacer(1, 1 * cm))
    story.append(Paragraph('Loss Documentation Report', title_style))
    story.append(Paragraph('FlockIQ — Farm Management Records', subtitle_style))
    story.append(Spacer(1, 0.6 * cm))

    cause = loss_details.get('cause_of_death', '')
    cause_label = _LOSS_CAUSE_LABELS.get(cause, cause.replace('_', ' ').title())
    incident_date = loss_details.get('incident_date')

    cover_data = [
        ['Farm', batch.farm.name],
        ['Batch', batch.batch_name],
        ['Incident Date', str(incident_date) if incident_date else 'N/A'],
        ['Report Generated', str(date.today())],
    ]
    ct = Table(cover_data, colWidths=[5 * cm, 11 * cm])
    ct.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 11),
        ('TEXTCOLOR', (0, 0), (0, -1), NAVY),
        ('ROWBACKGROUNDS', (0, 0), (-1, -1), [colors.white, STRIPE]),
        ('GRID', (0, 0), (-1, -1), 0.5, GRID),
        ('PADDING', (0, 0), (-1, -1), 8),
    ]))
    story.append(ct)
    story.append(Spacer(1, 0.7 * cm))

    # ── Section 1 — Incident Summary ──────────────────────────────────────────
    story.append(Paragraph('1. Incident Summary', section_style))
    incident_data = [
        ['Cause of Death', cause_label],
        ['Incident Date', str(incident_date) if incident_date else 'N/A'],
        ['Birds Affected', str(loss_details.get('birds_affected', 'N/A'))],
        ['Additional Notes', loss_details.get('additional_notes') or '—'],
    ]
    t1 = Table(incident_data, colWidths=[5 * cm, 11 * cm])
    t1.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('ROWBACKGROUNDS', (0, 0), (-1, -1), [colors.white, STRIPE]),
        ('GRID', (0, 0), (-1, -1), 0.5, GRID),
        ('PADDING', (0, 0), (-1, -1), 6),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    story.append(t1)
    story.append(Spacer(1, 0.5 * cm))

    # ── Section 2 — Flock Overview at Time of Loss ────────────────────────────
    story.append(Paragraph('2. Flock Overview at Time of Loss', section_style))
    overview_rows = [
        ['Bird Type', batch.get_bird_type_display()],
        ['Breed', batch.breed_name or '—'],
        ['Placement Date', str(batch.placement_date)],
        ['Initial Count', str(batch.initial_count)],
        ['Current Count', str(batch.current_count)],
    ]
    if valuation:
        conf = valuation.get('confidence', '')
        est = valuation.get('estimated_value_naira')
        val_text = f'NGN {est:,.0f}' if est is not None else 'N/A'
        if conf:
            val_text += f' ({conf} confidence)'
        overview_rows.append(['Estimated Value at Time of Loss', val_text])
    t2 = Table(overview_rows, colWidths=[6 * cm, 10 * cm])
    t2.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('ROWBACKGROUNDS', (0, 0), (-1, -1), [colors.white, STRIPE]),
        ('GRID', (0, 0), (-1, -1), 0.5, GRID),
        ('PADDING', (0, 0), (-1, -1), 6),
    ]))
    story.append(t2)
    if valuation and valuation.get('confidence') == 'low':
        story.append(Spacer(1, 0.15 * cm))
        story.append(Paragraph(
            'Note: estimated value is based on general market rates, not '
            'recorded sales.', styles['Normal'],
        ))
    story.append(Spacer(1, 0.5 * cm))

    # ── Section 3 — Vaccination & Health Compliance ───────────────────────────
    story.append(Paragraph('3. Vaccination &amp; Health Compliance', section_style))
    if vaccinations:
        vacc_data = [['Vaccine', 'Due Date', 'Status']]
        for v in vaccinations:
            status = v.get_status_display() if hasattr(v, 'get_status_display') else v.status
            vacc_data.append([v.vaccine_name, str(v.due_date), status])
        t3 = Table(vacc_data, colWidths=[8 * cm, 4 * cm, 4 * cm])
        t3.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), NAVY),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, STRIPE]),
            ('GRID', (0, 0), (-1, -1), 0.5, GRID),
            ('PADDING', (0, 0), (-1, -1), 5),
        ]))
        story.append(t3)
    else:
        story.append(Paragraph('No vaccination records on file for this batch.', styles['Normal']))
    story.append(Spacer(1, 0.5 * cm))

    # ── Section 4 — Mortality Timeline ────────────────────────────────────────
    story.append(Paragraph('4. Mortality Timeline', section_style))
    if mortality_logs:
        mort_data = [['Date', 'Count', 'Cause', 'Notes']]
        total_mort = 0
        for m in mortality_logs:
            total_mort += m.count
            cause_disp = m.get_cause_display() if hasattr(m, 'get_cause_display') else m.cause
            mort_data.append([str(m.date), str(m.count), cause_disp, (m.notes or '')[:60]])
        mort_data.append(['Total', str(total_mort), '', ''])
        t4 = Table(mort_data, colWidths=[3 * cm, 2 * cm, 3.5 * cm, 7.5 * cm])
        t4.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), NAVY),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('BACKGROUND', (0, -1), (-1, -1), STRIPE),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, STRIPE]),
            ('GRID', (0, 0), (-1, -1), 0.5, GRID),
            ('PADDING', (0, 0), (-1, -1), 5),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        story.append(t4)
    else:
        story.append(Paragraph('No mortality records on file for this batch.', styles['Normal']))
    story.append(Spacer(1, 0.5 * cm))

    # ── Section 5 — Feed & Care Log Summary ───────────────────────────────────
    story.append(Paragraph('5. Feed &amp; Care Log Summary', section_style))
    if feed_logs:
        total_feed = sum(float(f.quantity_kg) for f in feed_logs)
        dates = [f.record_date for f in feed_logs]
        first_date, last_date = min(dates), max(dates)
        span_days = (last_date - first_date).days + 1
        avg_daily = total_feed / span_days if span_days else total_feed
        feed_data = [
            ['Total Feed Consumed', f'{total_feed:,.1f} kg'],
            ['Average Daily Feed', f'{avg_daily:,.1f} kg/day'],
            ['Date Range', f'{first_date} to {last_date} ({span_days} days)'],
            ['Feed Records Logged', str(len(feed_logs))],
        ]
        t5 = Table(feed_data, colWidths=[6 * cm, 10 * cm])
        t5.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ROWBACKGROUNDS', (0, 0), (-1, -1), [colors.white, STRIPE]),
            ('GRID', (0, 0), (-1, -1), 0.5, GRID),
            ('PADDING', (0, 0), (-1, -1), 6),
        ]))
        story.append(t5)
    else:
        story.append(Paragraph('No feed records on file for this batch.', styles['Normal']))

    doc.build(story, onFirstPage=_loss_report_footer, onLaterPages=_loss_report_footer)
    return buffer.getvalue()


def generate_batch_report(batch) -> bytes:
    """Return a PDF byte string for a single batch summary."""
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=2 * cm,
        rightMargin=2 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
    )
    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph(f'Batch Report: {batch.batch_name}', styles['Title']))
    story.append(
        Paragraph(
            f'Farm: {batch.farm.name} | Generated: {date.today()}',
            styles['Normal'],
        )
    )
    story.append(Spacer(1, 0.5 * cm))

    data = [
        ['Field', 'Value'],
        ['Bird Type', batch.get_bird_type_display()],
        ['Placement Date', str(batch.placement_date)],
        ['Initial Count', str(batch.initial_count)],
        ['Current Count', str(batch.current_count)],
        ['Cycle Day', str(batch.cycle_day)],
        ['Mortality Rate', f'{batch.mortality_rate_pct:.1f}%'],
        ['Status', batch.get_status_display()],
    ]

    t = Table(data, colWidths=[6 * cm, 10 * cm])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), NAVY),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, STRIPE]),
        ('GRID', (0, 0), (-1, -1), 0.5, GRID),
        ('PADDING', (0, 0), (-1, -1), 6),
    ]))
    story.append(t)

    doc.build(story)
    return buffer.getvalue()


def generate_batch_excel(batch) -> bytes:
    """Return an Excel (.xlsx) byte string for a single batch summary."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Batch Summary'

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='3D5A99')

    for col, header in enumerate(['Field', 'Value'], 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    rows = [
        ('Batch Name', batch.batch_name),
        ('Bird Type', batch.get_bird_type_display()),
        ('Placement Date', str(batch.placement_date)),
        ('Initial Count', batch.initial_count),
        ('Current Count', batch.current_count),
        ('Cycle Day', batch.cycle_day),
        ('Mortality Rate', f'{batch.mortality_rate_pct:.1f}%'),
        ('Status', batch.get_status_display()),
    ]
    for row_idx, (field, value) in enumerate(rows, 2):
        ws.cell(row=row_idx, column=1, value=field)
        ws.cell(row=row_idx, column=2, value=value)

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 30

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def generate_production_overview_pdf(batch_summaries: list, today) -> bytes:
    """Return a PDF of the production overview across all active layer batches."""
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        leftMargin=2 * cm, rightMargin=2 * cm,
        topMargin=2 * cm, bottomMargin=2 * cm,
    )
    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph('Egg Production Overview', styles['Title']))
    story.append(Paragraph(f'Generated: {today}', styles['Normal']))
    story.append(Spacer(1, 0.5 * cm))

    data = [['Batch', 'Farm', 'Today\'s Eggs', 'Hen-Day %', '7-Day Avg %', 'Logged Today']]
    for s in batch_summaries:
        data.append([
            s['batch'].batch_name,
            s['batch'].farm.name,
            str(s['todays_eggs']),
            f"{s['todays_hen_day']:.1f}%",
            f"{s['avg_7day_hen_day']:.1f}%",
            'Yes' if s['logged_today'] else 'No',
        ])

    col_widths = [4 * cm, 4 * cm, 3 * cm, 3 * cm, 3 * cm, 3 * cm]
    t = Table(data, colWidths=col_widths)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), NAVY),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, STRIPE]),
        ('GRID', (0, 0), (-1, -1), 0.5, GRID),
        ('PADDING', (0, 0), (-1, -1), 5),
    ]))
    story.append(t)
    doc.build(story)
    return buffer.getvalue()


def generate_production_overview_excel(batch_summaries: list, today) -> bytes:
    """Return an Excel of the production overview."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Production Overview'

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='3D5A99')
    headers = ['Batch', 'Farm', "Today's Eggs", 'Hen-Day %', '7-Day Avg %', 'Logged Today']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    for row_idx, s in enumerate(batch_summaries, 2):
        ws.cell(row=row_idx, column=1, value=s['batch'].batch_name)
        ws.cell(row=row_idx, column=2, value=s['batch'].farm.name)
        ws.cell(row=row_idx, column=3, value=s['todays_eggs'])
        ws.cell(row=row_idx, column=4, value=round(s['todays_hen_day'], 1))
        ws.cell(row=row_idx, column=5, value=s['avg_7day_hen_day'])
        ws.cell(row=row_idx, column=6, value='Yes' if s['logged_today'] else 'No')

    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws.column_dimensions[col].width = 18

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def generate_vaccination_calendar_pdf(vaccinations: list, today) -> bytes:
    """Return a PDF of the vaccination calendar."""
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        leftMargin=2 * cm, rightMargin=2 * cm,
        topMargin=2 * cm, bottomMargin=2 * cm,
    )
    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph('Vaccination Calendar', styles['Title']))
    story.append(Paragraph(f'Generated: {today}', styles['Normal']))
    story.append(Spacer(1, 0.5 * cm))

    data = [['Batch', 'Farm', 'Vaccine', 'Due Date', 'Status', 'Days']]
    for v in vaccinations:
        data.append([
            v.batch.batch_name,
            v.batch.farm.name,
            v.vaccine_name,
            str(v.due_date),
            v.get_status_display() if hasattr(v, 'get_status_display') else v.status,
            getattr(v, 'days_label', ''),
        ])

    col_widths = [3.5 * cm, 3.5 * cm, 4 * cm, 3 * cm, 2.5 * cm, 2.5 * cm]
    t = Table(data, colWidths=col_widths)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), NAVY),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, STRIPE]),
        ('GRID', (0, 0), (-1, -1), 0.5, GRID),
        ('PADDING', (0, 0), (-1, -1), 5),
    ]))
    story.append(t)
    doc.build(story)
    return buffer.getvalue()


def generate_vaccination_calendar_excel(vaccinations: list, today) -> bytes:
    """Return an Excel of the vaccination calendar."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Vaccination Calendar'

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='3D5A99')
    headers = ['Batch', 'Farm', 'Vaccine', 'Due Date', 'Status', 'Days']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    for row_idx, v in enumerate(vaccinations, 2):
        ws.cell(row=row_idx, column=1, value=v.batch.batch_name)
        ws.cell(row=row_idx, column=2, value=v.batch.farm.name)
        ws.cell(row=row_idx, column=3, value=v.vaccine_name)
        ws.cell(row=row_idx, column=4, value=str(v.due_date))
        ws.cell(row=row_idx, column=5, value=v.get_status_display() if hasattr(v, 'get_status_display') else v.status)
        ws.cell(row=row_idx, column=6, value=getattr(v, 'days_label', ''))

    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws.column_dimensions[col].width = 18

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
