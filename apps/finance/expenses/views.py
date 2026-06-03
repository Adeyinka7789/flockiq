import csv
import datetime
import io
import json
from datetime import date, timedelta

import structlog
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Sum
from django.http import Http404, HttpResponse
from django.shortcuts import get_object_or_404, render
from django.views import View
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from apps.infrastructure.core.filters import DateRangeFilter
from apps.infrastructure.core.rls import set_tenant_context
from apps.infrastructure.core.views import TenantRequiredMixin

from .services import ExpenseService

logger = structlog.get_logger(__name__)


class ExpenseOverviewView(TenantRequiredMixin, View):
    """GET /expenses/ — farm-wide expense overview with break-even calculator."""

    def get(self, request):
        from apps.farm.farms.models import Farm
        from apps.farm.flocks.models import Batch

        from .models import ExpenseRecord

        org = request.user.org
        date_filter = DateRangeFilter()
        date_from, date_to = date_filter.get_date_range(request)
        preset = request.GET.get("preset", "30d")
        farm_id = request.GET.get("farm", "")
        batch_id = request.GET.get("batch", "")
        category = request.GET.get("category", "")

        # CSV export — resolve queryset then return early
        if request.GET.get("download") == "csv":
            with set_tenant_context(org):
                qs = (
                    ExpenseRecord.objects.filter(
                        expense_date__gte=date_from,
                        expense_date__lte=date_to,
                    )
                    .select_related("batch__farm")
                    .order_by("-expense_date")
                )
                if farm_id:
                    qs = qs.filter(farm_id=farm_id)
                if category:
                    qs = qs.filter(category=category)

                response = HttpResponse(content_type="text/csv")
                response["Content-Disposition"] = 'attachment; filename="expenses.csv"'
                writer = csv.writer(response)
                writer.writerow(["Date", "Description", "Category", "Batch", "Farm", "Amount (₦)"])
                for e in qs:
                    writer.writerow([
                        e.expense_date,
                        e.description,
                        e.category,
                        e.batch.batch_name if e.batch else "",
                        e.batch.farm.name if e.batch and e.batch.farm else "",
                        e.amount_kobo // 100,
                    ])
            return response

        with set_tenant_context(org):
            farms = list(Farm.objects.filter(is_active=True))

            qs = (
                ExpenseRecord.objects.filter(
                    expense_date__gte=date_from,
                    expense_date__lte=date_to,
                )
                .select_related("batch__farm")
                .order_by("-expense_date")
            )
            if farm_id:
                qs = qs.filter(farm_id=farm_id)
            if batch_id:
                qs = qs.filter(batch_id=batch_id)
            if category:
                qs = qs.filter(category=category)

            expenses = list(qs)
            total_kobo = sum(e.amount_kobo for e in expenses)
            total_naira = total_kobo // 100

            # Monthly trend — last 6 calendar months
            today = date.today()
            monthly_trend = []
            for i in range(5, -1, -1):
                month_start = (today.replace(day=1) - timedelta(days=i * 30)).replace(day=1)
                if month_start.month == 12:
                    next_month_start = month_start.replace(year=month_start.year + 1, month=1)
                else:
                    next_month_start = month_start.replace(month=month_start.month + 1)
                month_end = next_month_start - timedelta(days=1) if i > 0 else today

                month_qs = ExpenseRecord.objects.filter(
                    expense_date__gte=month_start,
                    expense_date__lte=month_end,
                )
                if farm_id:
                    month_qs = month_qs.filter(farm_id=farm_id)

                month_total = month_qs.aggregate(total=Sum("amount_kobo"))["total"] or 0
                monthly_trend.append({
                    "month": month_start.strftime("%b").upper(),
                    "total": month_total // 100,
                })

            # Category breakdown from current filtered expenses
            categories: dict = {}
            for e in expenses:
                cat = e.category or "other"
                categories[cat] = categories.get(cat, 0) + e.amount_kobo

            category_breakdown = [
                {"category": k, "total": v // 100}
                for k, v in sorted(categories.items(), key=lambda x: x[1], reverse=True)
            ]

            # Compare to previous period
            period_days = (date_to - date_from).days or 1
            prev_from = date_from - timedelta(days=period_days)
            prev_total = (
                ExpenseRecord.objects.filter(
                    expense_date__gte=prev_from,
                    expense_date__lt=date_from,
                ).aggregate(total=Sum("amount_kobo"))["total"]
                or 0
            )
            if prev_total > 0:
                change_pct = round((total_kobo - prev_total) / prev_total * 100, 1)
            else:
                change_pct = None

            batches = list(Batch.objects.filter(status="active").select_related("farm"))

            # Break-even data for up to 5 active batches
            breakeven_batches = []
            for batch in batches[:5]:
                batch_expense = (
                    ExpenseRecord.objects.filter(batch=batch).aggregate(total=Sum("amount_kobo"))["total"]
                    or 0
                )
                breakeven_batches.append({
                    "batch": batch,
                    "total_expenses_naira": batch_expense // 100,
                })

        context = {
            "expenses": expenses,
            "total_naira": total_naira,
            "change_pct": change_pct,
            "monthly_trend": monthly_trend,
            "category_breakdown": category_breakdown,
            "farms": farms,
            "batches": batches,
            "active_farm": farm_id,
            "active_batch": batch_id,
            "active_category": category,
            "active_preset": preset,
            "date_from": date_from.strftime("%Y-%m-%d"),
            "date_to": date_to.strftime("%Y-%m-%d"),
            "category_choices": ExpenseRecord.CATEGORY_CHOICES,
            "breakeven_batches": breakeven_batches,
            "today": date.today(),
        }

        if request.headers.get("HX-Request"):
            return render(request, "finance/expenses/_expense_overview_partial.html", context)
        return render(request, "finance/expenses/expense_overview.html", context)


def _get_org(request):
    org = getattr(request.user, "org", None)
    if org is None:
        raise Http404("No organisation found for this user.")
    return org


class ExpenseLogView(LoginRequiredMixin, View):
    """GET+POST /finance/expenses/<uuid:batch_pk>/log/"""

    def get(self, request, batch_pk):
        from apps.farm.flocks.models import Batch
        from django.shortcuts import get_object_or_404
        from .models import ExpenseRecord

        org = _get_org(request)
        with set_tenant_context(org):
            batch = get_object_or_404(Batch, pk=batch_pk, org=org)
        return render(request, "expenses/_expense_log_form.html", {
            "batch": batch,
            "batch_pk": batch_pk,
            "today": datetime.date.today(),
            "categories": ExpenseRecord.CATEGORY_CHOICES,
        })

    def post(self, request, batch_pk):
        from .models import ExpenseRecord

        org = _get_org(request)
        data = request.POST

        def _error(msg, status=422):
            return render(request, "expenses/_expense_log_form.html", {
                "error": msg,
                "batch_pk": batch_pk,
                "today": datetime.date.today(),
                "categories": ExpenseRecord.CATEGORY_CHOICES,
            }, status=status)

        for field in ["category", "amount", "description"]:
            if not data.get(field):
                return _error(f"{field} is required.")

        try:
            amount_naira = float(data["amount"])
            amount_kobo = int(amount_naira * 100)
        except (ValueError, TypeError):
            return _error("Invalid amount.")

        expense_date_str = data.get("expense_date")
        expense_date = None
        if expense_date_str:
            try:
                expense_date = datetime.date.fromisoformat(expense_date_str)
            except ValueError:
                pass

        farm_id = data.get("farm_id", "")

        with set_tenant_context(org):
            try:
                record = ExpenseService(org).record_expense(
                    farm_id=farm_id,
                    category=data["category"],
                    amount_kobo=amount_kobo,
                    description=data["description"],
                    expense_date=expense_date,
                    batch_id=str(batch_pk),
                    receipt_ref=data.get("receipt_ref", ""),
                    notes=data.get("notes", ""),
                    recorded_by=request.user,
                )
            except ValueError as exc:
                return _error(str(exc))

        response = render(
            request,
            "expenses/_expense_log_form.html",
            {"success": True, "record": record, "batch_pk": batch_pk},
        )
        response["HX-Trigger"] = '{"expenseAdded": true}'
        return response


class ExpenseTableView(LoginRequiredMixin, View):
    """GET /finance/expenses/<uuid:batch_pk>/table/ — supports date range + category filters."""

    def get(self, request, batch_pk):
        from apps.infrastructure.core.filters import DateRangeFilter
        from .models import ExpenseRecord

        org = _get_org(request)
        df = DateRangeFilter()
        date_from, date_to = df.get_date_range(request)
        filter_ctx = df.get_filter_context(request)
        category = request.GET.get("category", "")

        with set_tenant_context(org):
            expenses = ExpenseService(org).get_batch_expenses(str(batch_pk))
            expenses = expenses.filter(
                expense_date__gte=date_from,
                expense_date__lte=date_to,
            )
            if category:
                expenses = expenses.filter(category=category)
            expenses = list(expenses)

        PRESETS = [
            ("7d", "7 days"),
            ("30d", "30 days"),
            ("90d", "90 days"),
            ("this_month", "This month"),
        ]

        return render(
            request,
            "expenses/_expense_table.html",
            {
                "expenses": expenses,
                "batch_pk": batch_pk,
                "active_category": category,
                "category_choices": ExpenseRecord.CATEGORY_CHOICES,
                "presets": PRESETS,
                **filter_ctx,
            },
        )


class ExpenseBreakdownView(LoginRequiredMixin, View):
    """GET /finance/expenses/<uuid:batch_pk>/breakdown/"""

    def get(self, request, batch_pk):
        org = _get_org(request)
        with set_tenant_context(org):
            breakdown = ExpenseService(org).get_expense_breakdown(batch_id=str(batch_pk))
        return render(
            request,
            "expenses/_expense_breakdown_chart.html",
            {"breakdown": breakdown, "batch_pk": batch_pk},
        )


class ExpenseFarmSummaryView(LoginRequiredMixin, View):
    """GET /finance/expenses/farm/<uuid:farm_pk>/summary/"""

    def get(self, request, farm_pk):
        org = _get_org(request)
        month = request.GET.get("month")
        with set_tenant_context(org):
            summary = ExpenseService(org).get_farm_expenses_summary(
                farm_id=str(farm_pk),
                month=int(month) if month else None,
            )
        return render(
            request,
            "expenses/_farm_expense_summary.html",
            {"summary": summary, "farm_pk": farm_pk},
        )


class ExpenseAPIView(APIView):
    """GET+POST /api/v1/expenses/"""

    permission_classes = [IsAuthenticated]

    def get(self, request):
        org = _get_org(request)
        batch_id = request.query_params.get("batch_id")
        with set_tenant_context(org):
            svc = ExpenseService(org)
            if batch_id:
                expenses = svc.get_batch_expenses(batch_id)
                data = [
                    {
                        "id": str(e.id),
                        "category": e.category,
                        "amount_naira": e.amount_naira,
                        "description": e.description,
                        "expense_date": e.expense_date.isoformat(),
                    }
                    for e in expenses
                ]
            else:
                data = []
        return Response({"data": data})

    def post(self, request):
        org = _get_org(request)
        d = request.data

        try:
            amount_kobo = int(float(d.get("amount_naira", 0)) * 100)
            with set_tenant_context(org):
                record = ExpenseService(org).record_expense(
                    farm_id=d.get("farm_id", ""),
                    category=d.get("category", ""),
                    amount_kobo=amount_kobo,
                    description=d.get("description", ""),
                    batch_id=d.get("batch_id"),
                    receipt_ref=d.get("receipt_ref", ""),
                    notes=d.get("notes", ""),
                    recorded_by=request.user,
                )
        except ValueError as exc:
            return Response({"error": str(exc)}, status=400)

        return Response(
            {
                "id": str(record.id),
                "category": record.category,
                "amount_naira": record.amount_naira,
            },
            status=201,
        )


class FinancePDFExportView(TenantRequiredMixin, View):
    def get(self, request, batch_pk):
        from apps.infrastructure.billing.features import has_feature

        if not has_feature(request.user.org, "pdf_export"):
            response = HttpResponse(status=200)
            response["HX-Trigger"] = json.dumps({
                "showToast": {
                    "message": "PDF exports are on the Monthly plan. Upgrade to unlock.",
                    "type": "error",
                }
            })
            return response

        from apps.farm.flocks.models import Batch
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import (
            Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle)

        from apps.finance.finance.models import SalesRecord

        from .models import ExpenseRecord

        with set_tenant_context(request.user.org):
            batch = get_object_or_404(Batch, pk=batch_pk)
            expenses = list(ExpenseRecord.objects.filter(batch=batch).order_by("-expense_date"))
            sales = list(SalesRecord.objects.filter(batch=batch).order_by("-sale_date"))
            total_revenue = sum(s.total_amount_kobo for s in sales) // 100
            total_expenses = sum(e.amount_kobo for e in expenses) // 100
            profit = total_revenue - total_expenses

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []

        story.append(Paragraph(f"Financial Report — {batch.batch_name}", styles["Title"]))
        story.append(Paragraph(
            f"{batch.farm.name} | {batch.bird_type.title()} | Day {batch.cycle_day}",
            styles["Normal"],
        ))
        story.append(Spacer(1, 20))

        story.append(Paragraph("P&L Summary", styles["Heading2"]))
        pl_data = [
            ["", "Amount"],
            ["Total Revenue", f"₦{total_revenue:,}"],
            ["Total Expenses", f"₦{total_expenses:,}"],
            ["Net Profit", f"₦{profit:,}"],
        ]
        pl_table = Table(pl_data, colWidths=[300, 150])
        pl_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#3d5a99")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f4f7fe")]),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e2e8f0")),
            ("PADDING", (0, 0), (-1, -1), 8),
        ]))
        story.append(pl_table)
        story.append(Spacer(1, 20))

        if sales:
            story.append(Paragraph("Sales Records", styles["Heading2"]))
            sales_data = [["Date", "Product", "Qty", "Unit Price", "Total"]]
            for s in sales:
                sales_data.append([
                    str(s.sale_date),
                    s.product_type,
                    str(s.quantity),
                    f"₦{s.unit_price_kobo // 100:,}",
                    f"₦{s.total_amount_kobo // 100:,}",
                ])
            sales_table = Table(sales_data, colWidths=[80, 100, 60, 100, 100])
            sales_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#3d5a99")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f4f7fe")]),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e2e8f0")),
                ("PADDING", (0, 0), (-1, -1), 6),
            ]))
            story.append(sales_table)
            story.append(Spacer(1, 20))

        if expenses:
            story.append(Paragraph("Expense Records", styles["Heading2"]))
            exp_data = [["Date", "Category", "Description", "Amount"]]
            for e in expenses:
                exp_data.append([
                    str(e.expense_date),
                    e.category,
                    e.description[:40],
                    f"₦{e.amount_kobo // 100:,}",
                ])
            exp_table = Table(exp_data, colWidths=[80, 100, 180, 100])
            exp_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#3d5a99")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f4f7fe")]),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e2e8f0")),
                ("PADDING", (0, 0), (-1, -1), 6),
            ]))
            story.append(exp_table)

        doc.build(story)
        buffer.seek(0)

        response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
        response["Content-Disposition"] = (
            f'attachment; filename="finance_{batch.batch_name}.pdf"'
        )
        return response


class FinanceExcelExportView(TenantRequiredMixin, View):
    def get(self, request, batch_pk):
        from apps.infrastructure.billing.features import has_feature

        if not has_feature(request.user.org, "excel_export"):
            response = HttpResponse(status=200)
            response["HX-Trigger"] = json.dumps({
                "showToast": {
                    "message": "Excel exports are on the Monthly plan. Upgrade to unlock.",
                    "type": "error",
                }
            })
            return response

        import openpyxl
        from openpyxl.styles import Font, PatternFill

        from apps.farm.flocks.models import Batch
        from apps.finance.finance.models import SalesRecord

        from .models import ExpenseRecord

        with set_tenant_context(request.user.org):
            batch = get_object_or_404(Batch, pk=batch_pk)
            expenses = list(ExpenseRecord.objects.filter(batch=batch).order_by("-expense_date"))
            sales = list(SalesRecord.objects.filter(batch=batch).order_by("-sale_date"))

        wb = openpyxl.Workbook()
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(fill_type="solid", fgColor="3d5a99")

        ws_sales = wb.active
        ws_sales.title = "Sales"
        ws_sales.append(["Date", "Product", "Quantity", "Unit Price (₦)", "Total (₦)"])
        for cell in ws_sales[1]:
            cell.font = header_font
            cell.fill = header_fill
        for s in sales:
            ws_sales.append([
                str(s.sale_date),
                s.product_type,
                s.quantity,
                s.unit_price_kobo // 100,
                s.total_amount_kobo // 100,
            ])

        ws_exp = wb.create_sheet("Expenses")
        ws_exp.append(["Date", "Category", "Description", "Amount (₦)"])
        for cell in ws_exp[1]:
            cell.font = header_font
            cell.fill = header_fill
        for e in expenses:
            ws_exp.append([
                str(e.expense_date),
                e.category,
                e.description,
                e.amount_kobo // 100,
            ])

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = (
            f'attachment; filename="finance_{batch.batch_name}.xlsx"'
        )
        return response
